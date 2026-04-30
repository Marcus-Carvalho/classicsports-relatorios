# -*- coding: utf-8 -*-
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import asyncio
import shutil
import tempfile
import threading
import ctypes as _ctypes
from pathlib import Path
import pandas as pd
from playwright.async_api import async_playwright

# ── CONFIGURACAO ──────────────────────────────────────────────
# Importa do config.py que fica na mesma pasta
import importlib.util, os
_cfg_path = Path(__file__).parent / "config.py"
_spec = importlib.util.spec_from_file_location("config", _cfg_path)
_cfg  = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_cfg)

# Detecta Chrome Profile automaticamente na maquina do usuario
import os as _os2
_chrome_auto = _os2.path.join(_os2.path.expanduser('~'), 'AppData', 'Local', 'Google', 'Chrome', 'User Data')
CHROME_ORIGINAL = getattr(_cfg, 'CHROME_PROFILE_DIR', _chrome_auto) if _cfg else _chrome_auto
PERFIL          = "Profile 13"
LOGIN_URL       = "https://app.sportbayhub.com.br/login"
URL_ANUNCIOS    = "https://app.sportbayhub.com.br/produtos"
URL_MLB_SKU     = "https://app.sportbayhub.com.br/produtos/upload_id"
EMAIL           = "classic.sports.full.11@gmail.com"
SENHA           = "CM86EFh.PAbxqPr"
NOME_LOJA       = "11"
NOME_ARQUIVO    = "11"

PASTA_SAIDA = Path(__file__).parent / f"relatorios_{NOME_ARQUIVO}"
PASTA_SAIDA.mkdir(exist_ok=True)

# ── POPUP ─────────────────────────────────────────────────────
def _popup(msg):
    _ctypes.windll.user32.MessageBoxW(0, msg, "SportBay Automacao", 1)


# ── COPIA PERFIL ──────────────────────────────────────────────
def copiar_perfil():
    tmp = Path(tempfile.mkdtemp(prefix="chrome_sportbay_"))
    origem  = Path(CHROME_ORIGINAL) / PERFIL
    destino = tmp / "Default"
    print("  [COPIANDO] Copiando perfil...")

    def copiar_ignorando_erros(src, dst, ignore=None):
        os.makedirs(dst, exist_ok=True)
        for item in os.listdir(src):
            s = os.path.join(src, item)
            d = os.path.join(dst, item)
            if ignore and item in ignore(src, os.listdir(src)):
                continue
            try:
                if os.path.isdir(s):
                    copiar_ignorando_erros(s, d, ignore)
                else:
                    shutil.copy2(s, d)
            except Exception:
                pass

    ignorar = shutil.ignore_patterns(
        "*.log", "*.tmp", "GPUCache", "ShaderCache", "Code Cache",
        "Cache", "CachedData", "DawnCache", "Cookies*", "Sessions", "Safe Browsing*"
    )
    copiar_ignorando_erros(str(origem), str(destino), ignorar)
    print("  [OK] Perfil copiado!")
    return tmp


# ── PROCV: SKU + CUSTO + MARGEM ───────────────────────────────
def fazer_procv_completo():
    arq_anuncios  = PASTA_SAIDA / "listas_de_anuncios.xlsx"
    arq_mlb_sku   = PASTA_SAIDA / "mlb_id_interno.xlsx"
    arq_resultado = PASTA_SAIDA / "listas_de_anuncios_COM_SKU_CUSTO_MARGEM.xlsx"
    pasta_scripts = Path(__file__).parent

    # Localiza arquivo de tabela — busca em múltiplas pastas com fallback inteligente
    def localizar_arquivo(pasta, nome_cfg):
        if not nome_cfg:
            return pasta / "nao_configurado.xlsx"

        # 1. Nome exato na pasta do script (tmp_usuario)
        arq = pasta / nome_cfg
        if arq.exists():
            return arq

        # 2. Nome exato na pasta PAI (raiz de instalação: C:\ClassicSportsApps\Relatorios)
        arq_pai = pasta.parent / nome_cfg
        if arq_pai.exists():
            print(f"  [INFO] Tabela encontrada na pasta pai: {arq_pai}")
            return arq_pai

        # 3. Busca por similaridade em ambas as pastas
        palavras = [p for p in nome_cfg.lower().replace("-","_").replace(" ","_")
                    .replace(".xlsx","").split("_") if len(p) > 3]

        for busca_pasta in [pasta, pasta.parent]:
            for xlsx in busca_pasta.glob("*.xlsx"):
                nome_f = xlsx.name.lower().replace("-","_").replace(" ","_")
                matches = sum(1 for p in palavras if p in nome_f)
                if matches >= 2 or (len(palavras) == 1 and palavras[0] in nome_f):
                    print(f"  [INFO] Usando '{xlsx.name}' no lugar de '{nome_cfg}'")
                    return xlsx

        # 4. Busca por tipo de tabela (precos/kits) independente do nome
        tipo_map = {
            "precos": ["preco", "price", "custo", "tabela"],
            "kits": ["kit", "meus_kit", "composicao"],
            "sku_kits": ["sku", "kit", "preco_sku"],
        }
        nome_lower = nome_cfg.lower()
        tipo = None
        if "meus" in nome_lower and "kit" in nome_lower:
            tipo = "kits"
        elif "sku" in nome_lower and "kit" in nome_lower:
            tipo = "sku_kits"
        elif "preco" in nome_lower or "tabela" in nome_lower:
            tipo = "precos"

        if tipo:
            palavras_tipo = tipo_map[tipo]
            for busca_pasta in [pasta, pasta.parent]:
                candidatos = []
                for xlsx in busca_pasta.glob("*.xlsx"):
                    nome_f = xlsx.name.lower().replace("-","_")
                    score = sum(1 for p in palavras_tipo if p in nome_f)
                    if score >= 1:
                        candidatos.append((score, xlsx))
                if candidatos:
                    candidatos.sort(key=lambda x: -x[0])
                    melhor = candidatos[0][1]
                    print(f"  [INFO] Melhor match para '{nome_cfg}': '{melhor.name}'")
                    return melhor

        print(f"  [AVISO] Tabela nao encontrada: '{nome_cfg}'")
        return arq  # retorna original (não existe — será avisado depois)

    arq_tabela_precos  = localizar_arquivo(pasta_scripts, _cfg.TABELA_PRECOS)
    arq_preco_sku_kits = localizar_arquivo(pasta_scripts, _cfg.TABELA_PRECO_SKU_KITS)
    arq_meus_kits      = localizar_arquivo(pasta_scripts, _cfg.TABELA_MEUS_KITS)

    if not arq_anuncios.exists() or not arq_mlb_sku.exists():
        print("  [AVISO] Arquivos nao encontrados para cruzamento.")
        return

    tabelas_faltando = []
    for arq, nome in [(arq_tabela_precos, _cfg.TABELA_PRECOS),
                      (arq_preco_sku_kits, _cfg.TABELA_PRECO_SKU_KITS),
                      (arq_meus_kits, _cfg.TABELA_MEUS_KITS)]:
        if not arq.exists():
            tabelas_faltando.append(nome)
            print(f"  [AVISO] Tabela nao encontrada: {nome}")
    if tabelas_faltando:
        msg = "AVISO: Tabelas nao encontradas (PROCV parcial):\n"
        msg += "\n".join(f"- {t}" for t in tabelas_faltando)
        msg += "\n\nO processo continuara com as tabelas disponiveis."
        print(f"  [AVISO] Continuando com tabelas parciais: {tabelas_faltando}")
        _popup(msg)

    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    df_anuncios  = pd.read_excel(arq_anuncios)
    df_sku_tab   = pd.read_excel(arq_mlb_sku)
    # Detecta se a tabela de precos tem linha de titulo antes do cabecalho
    def ler_tabela_precos(arq):
        if not arq.exists():
            return pd.DataFrame()
        # Tenta sem skiprows primeiro — se tiver coluna SKU, esta correto
        df_test = pd.read_excel(arq)
        colunas_norm = [str(c).lower().replace(" ","").replace("_","") for c in df_test.columns]
        if any("sku" in c for c in colunas_norm):
            return df_test
        # Tenta com skiprows=1
        df_skip = pd.read_excel(arq, skiprows=1)
        colunas_norm2 = [str(c).lower().replace(" ","").replace("_","") for c in df_skip.columns]
        if any("sku" in c for c in colunas_norm2):
            return df_skip
        # Tenta skiprows=0 a 3
        for skip in range(0, 4):
            df_try = pd.read_excel(arq, skiprows=skip)
            cols = [str(c).lower() for c in df_try.columns]
            if any("sku" in c for c in cols):
                return df_try
        return df_test  # retorna sem skiprows como fallback
    df_precos    = ler_tabela_precos(arq_tabela_precos)
    df_kits      = pd.read_excel(arq_preco_sku_kits)            if arq_preco_sku_kits.exists() else pd.DataFrame()
    df_meus_kits = pd.read_excel(arq_meus_kits)                 if arq_meus_kits.exists()      else pd.DataFrame()

    # PROCV 1: MLB -> SKU
    col_mlb_an  = detectar_coluna(df_anuncios, ["mlb","id do anuncio","item_id","anuncio"])
    col_mlb_sku = detectar_coluna(df_sku_tab,  ["mlb","id do anuncio","item_id","anuncio"])
    col_id_int  = detectar_coluna(df_sku_tab,  ["id interno","sku","cod","codigo"])
    if col_mlb_an and col_mlb_sku and col_id_int:
        lookup_sku = df_sku_tab.set_index(col_mlb_sku)[col_id_int].to_dict()
        df_anuncios["SKU"] = df_anuncios[col_mlb_an].map(lookup_sku)
        print(f"  [OK] SKU: {df_anuncios['SKU'].notna().sum()}/{len(df_anuncios)} encontrados")

    # Monta lookups
    lookup_sku_custo = lookup_sku_margem = {}
    lookup_titulo_custo = custo_kit_sis = {}

    if not df_precos.empty and "SKU" in df_precos.columns:
        df_precos["_SK"] = df_precos["SKU"].str.strip().str.upper()
        col_cg = detectar_coluna(df_precos, ["custo geral","custo_geral"])
        col_mg = detectar_coluna(df_precos, ["margem minima","margem_minima","margem min","margem"])
        if col_cg: lookup_sku_custo  = df_precos.set_index("_SK")[col_cg].to_dict()
        if col_mg: lookup_sku_margem = df_precos.set_index("_SK")[col_mg].to_dict()

    if not df_kits.empty and "Produtos" in df_kits.columns and "CUSTO" in df_kits.columns:
        lookup_titulo_custo = df_kits.set_index(
            df_kits["Produtos"].str.strip().str.lower()
        )["CUSTO"].to_dict()

    if not df_meus_kits.empty:
        col_sk = detectar_coluna(df_meus_kits, ["sku_kit","sku kit"])
        col_pr = detectar_coluna(df_meus_kits, ["preco_sportbay","preco sportbay","sportbay"])
        col_qt = detectar_coluna(df_meus_kits, ["quantidade","qtd"])
        if col_sk and col_pr and col_qt:
            custo_kit_sis = df_meus_kits.groupby(col_sk).apply(
                lambda x: (x[col_pr] * x[col_qt]).sum()
            ).to_dict()

    def buscar_titulo_sem_sufixo(titulo):
        palavras = titulo.strip().lower().split()
        for n in range(1, 4):
            t = " ".join(palavras[:-n])
            if t in lookup_titulo_custo:
                return lookup_titulo_custo[t]
        return None

    def buscar_custo(row):
        titulo = str(row.get("Titulo", row.get("Título", ""))).strip().lower()
        sku    = str(row.get("SKU","")).strip().upper() if pd.notna(row.get("SKU")) else ""
        if sku   in lookup_sku_custo:    return lookup_sku_custo[sku]
        if titulo in lookup_titulo_custo: return lookup_titulo_custo[titulo]
        c = buscar_titulo_sem_sufixo(titulo)
        if c is not None:                return c
        if sku in custo_kit_sis:         return custo_kit_sis[sku]
        return None

    MARGEM_PADRAO = getattr(_cfg, "MARGEM_PADRAO", 16)

    def buscar_margem(row):
        sku = str(row.get("SKU","")).strip().upper() if pd.notna(row.get("SKU")) else ""
        # Busca margem por SKU na tabela de preços
        if sku and sku in lookup_sku_margem:
            val = lookup_sku_margem.get(sku)
            if pd.notna(val):
                try:
                    return float(val)
                except (ValueError, TypeError):
                    pass
        # SKU não encontrado na tabela — usa margem padrão
        return MARGEM_PADRAO

    df_anuncios["Custo"]             = df_anuncios.apply(buscar_custo, axis=1)
    df_anuncios["Margem Minima (%)"] = df_anuncios.apply(buscar_margem, axis=1)

    # Catalogo: busca em outros anuncios da propria planilha
    col_titulo = "Titulo" if "Titulo" in df_anuncios.columns else "Título"
    df_com = df_anuncios[df_anuncios["Custo"].notna()].copy()
    df_com["_tl"] = df_com[col_titulo].str.strip().str.lower()
    li_c = df_com.groupby("_tl")["Custo"].first().to_dict()
    li_m = df_com.groupby("_tl")["Margem Minima (%)"].first().to_dict()
    for idx in df_anuncios[df_anuncios["Custo"].isna()].index:
        tl = str(df_anuncios.at[idx, col_titulo]).strip().lower().split()
        for n in range(1, 4):
            base = " ".join(tl[:-n])
            for t in li_c:
                if t.startswith(base) and len(t) > len(base):
                    df_anuncios.at[idx, "Custo"] = li_c[t]
                    df_anuncios.at[idx, "Margem Minima (%)"] = li_m.get(t, 16)
                    break
            else: continue
            break

    print(f"  [OK] Custo: {df_anuncios['Custo'].notna().sum()}/{len(df_anuncios)} encontrados")

    # Reordena colunas
    cols = list(df_anuncios.columns)
    for c in ["Custo", "Margem Minima (%)"]:
        if c in cols: cols.remove(c)
    if "SKU" in cols:
        i = cols.index("SKU") + 1
        cols.insert(i,   "Custo")
        cols.insert(i+1, "Margem Minima (%)")
    df_anuncios = df_anuncios[cols]

    df_anuncios.to_excel(arq_resultado, index=False, sheet_name="Anuncios")
    wb = openpyxl.load_workbook(arq_resultado)
    ws = wb["Anuncios"]
    col_map = {ws.cell(1,c).value: c for c in range(1, ws.max_column+1)}

    # Insere Preco Minimo
    ci_marg = col_map["Margem Minima (%)"]
    ws.insert_cols(ci_marg + 1)
    ws.cell(1, ci_marg + 1).value = "Preco Minimo (R$)"
    col_map = {ws.cell(1,c).value: c for c in range(1, ws.max_column+1)}

    ci_frete = col_map.get("Tarifa Frete Grátis")
    ci_taxa  = col_map.get("Porcentagem cobrada do ML")
    ci_custo = col_map.get("Custo")
    ci_marg  = col_map.get("Margem Minima (%)")
    ci_preco = col_map.get("Preco Minimo (R$)")
    IMPOSTO  = 0.17

    for row in range(2, ws.max_row + 1):
        custo  = ws.cell(row, ci_custo).value if ci_custo else None
        frete  = ws.cell(row, ci_frete).value if ci_frete else 0
        taxa   = ws.cell(row, ci_taxa).value  if ci_taxa  else None
        margem = ws.cell(row, ci_marg).value  if ci_marg  else None
        frete  = frete if isinstance(frete,(int,float)) and frete > 0 else 0
        if isinstance(custo,(int,float)) and isinstance(taxa,(int,float)) and isinstance(margem,(int,float)):
            div = 1 - (margem/100) - IMPOSTO - (taxa/100)
            ws.cell(row, ci_preco).value = round((custo+frete)/div, 2) if div > 0 else "Indefinido"
        else:
            ws.cell(row, ci_preco).value = "Indefinido"

    FILLS = {
        "SKU":               PatternFill("solid", start_color="D6EAF8"),
        "Custo":             PatternFill("solid", start_color="D5F5E3"),
        "Margem Minima (%)": PatternFill("solid", start_color="FEF9E7"),
        "Preco Minimo (R$)": PatternFill("solid", start_color="FADBD8"),
    }
    H_N = PatternFill("solid", start_color="2C3E50")
    H_D = PatternFill("solid", start_color="1A5276")
    HF  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    HA  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    col_map = {ws.cell(1,c).value: c for c in range(1, ws.max_column+1)}

    for col in range(1, ws.max_column+1):
        cell = ws.cell(1, col)
        cell.fill = H_D if cell.value in FILLS else H_N
        cell.font = HF
        cell.alignment = HA
    ws.row_dimensions[1].height = 35

    for cn, fill in FILLS.items():
        ci = col_map.get(cn)
        if not ci: continue
        for row in range(2, ws.max_row+1):
            cell = ws.cell(row, ci)
            cell.fill = fill
            cell.font = Font(name="Arial", size=9)
            if cn in ("Custo","Preco Minimo (R$)") and isinstance(cell.value,(int,float)):
                cell.number_format = "R$ #,##0.00"
            elif cn == "Margem Minima (%)" and isinstance(cell.value,(int,float)):
                cell.number_format = "0.0\"%\""
            if cell.value == "Indefinido":
                cell.font = Font(name="Arial", size=9, color="C0392B", bold=True)

    LG = {"MLB":22,"SKU":18,"Custo":14,"Margem Minima (%)":16,"Preco Minimo (R$)":16,
          "Título":55,"Apelido":20,"Status":12,"Preço Por":13,"Preço De":13,
          "Porcentagem cobrada do ML":14,"Tarifa Frete Grátis":14}
    for col in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(col)].width = LG.get(ws.cell(1,col).value, 13)

    ws.freeze_panes = "A2"
    wb.save(arq_resultado)
    print("  [OK] Arquivo final salvo: listas_de_anuncios_COM_SKU_CUSTO_MARGEM.xlsx")

def detectar_coluna(df, palavras_chave):
    """Detecta coluna ignorando acentos, espaços, underscores, parênteses e %."""
    import unicodedata
    def _norm(s):
        s = str(s).lower().strip()
        s = "".join(c for c in unicodedata.normalize("NFD", s)
                    if unicodedata.category(c) != "Mn")
        return s.replace(" ","").replace("_","").replace("(","").replace(")","").replace("%","").replace(".","")
    for col in df.columns:
        col_n = _norm(col)
        for palavra in palavras_chave:
            if _norm(palavra) in col_n:
                return col
    return None


# ── AUTOMACAO PRINCIPAL ───────────────────────────────────────
async def main():
    print(f"\n[INICIO] Iniciando automacao - Loja: {NOME_LOJA}\n")
    for f in PASTA_SAIDA.glob("*.xlsx"):
        try:
            f.unlink()
        except Exception:
            pass

    tmp_dir = copiar_perfil()

    try:
        async with async_playwright() as p:
            context = await p.chromium.launch_persistent_context(
                user_data_dir=str(tmp_dir),
                channel="chrome",  # Usa Chrome instalado do usuario
                headless=False,
                accept_downloads=True,
                args=["--disable-blink-features=AutomationControlled", "--no-sandbox", "--disable-infobars"],
                ignore_default_args=["--enable-automation"],
            )
            await context.add_init_script("Object.defineProperty(navigator,'webdriver',{{get:()=>undefined}})")
            page = context.pages[0] if context.pages else await context.new_page()

            # LOGIN
            print("[LOGIN] Abrindo SportBay Hub...")
            try:
                await page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=30000)
            except Exception:
                pass  # Continua mesmo se timeout no goto
            await page.wait_for_timeout(2000)

            if "login" in page.url.lower():
                print("  Fazendo login...")
                try:
                    await page.click("button:has-text('Continuar')", timeout=3000)
                    await page.wait_for_timeout(800)
                except Exception:
                    pass

                await page.wait_for_selector("input[placeholder='Email']", timeout=10000)
                await page.fill("input[placeholder='Email']", "")
                await page.type("input[placeholder='Email']", EMAIL, delay=80)
                await page.wait_for_timeout(400)
                await page.fill("input[placeholder='Senha']", "")
                await page.type("input[placeholder='Senha']", SENHA, delay=80)
                await page.wait_for_timeout(400)

                popup_done = threading.Event()
                def show_popup():
                    _popup("Verifique se apareceu CAPTCHA no navegador.\nSe sim, resolva-o primeiro.\nDepois clique OK para fazer o login.")
                    popup_done.set()
                t = threading.Thread(target=show_popup)
                t.start()
                while not popup_done.is_set():
                    await page.wait_for_timeout(500)

                try:
                    await page.click("button:has-text('Login')", timeout=10000)
                except Exception:
                    pass
                await page.wait_for_timeout(3000)

                # Aguarda o usuario completar todo o login (incluindo MFA se necessario)
                print("  Aguardando conclusao do login (incluindo MFA se necessario)...")
                mfa_avisado = False
                max_espera  = 180  # 3 minutos maximo
                for _ in range(max_espera):
                    url_atual = page.url.lower()
                    # Saiu do login = sucesso
                    if "login" not in url_atual:
                        print("  [OK] Login concluido! URL: " + page.url)
                        break
                    # Detecta MFA pela presenca do campo de codigo ou botao Validar
                    try:
                        tem_mfa = await page.evaluate("""
                            () => {
                                var texto = document.body.innerText || '';
                                return texto.includes('Validar') ||
                                       texto.includes('MFA') ||
                                       texto.includes('Autenticac') ||
                                       texto.includes('Verificac') ||
                                       texto.includes('Escolha um m') ||
                                       texto.includes('codigo') ||
                                       texto.includes('código');
                            }
                        """)
                        if tem_mfa and not mfa_avisado:
                            mfa_avisado = True
                            print("  [MFA] Autenticacao em 2 fatores detectada!")
                            popup_done_mfa = threading.Event()
                            def show_popup_mfa():
                                _popup("Autenticacao em 2 fatores detectada!\n\n1. Escolha o metodo no navegador\n2. Digite o codigo recebido\n3. Clique em Validar Codigo\n\nDepois clique OK aqui.")
                                popup_done_mfa.set()
                            t_mfa = threading.Thread(target=show_popup_mfa)
                            t_mfa.start()
                            while not popup_done_mfa.is_set():
                                await page.wait_for_timeout(500)
                            print("  [MFA] Usuario confirmou - aguardando redirecionamento...")
                            # Aguarda ate 60s para sair do login apos MFA
                            for _w in range(60):
                                if "login" not in page.url.lower():
                                    break
                                await page.wait_for_timeout(1000)
                            mfa_avisado = False
                    except Exception:
                        pass
                    await page.wait_for_timeout(1000)
                else:
                    print("  [ERRO] Tempo esgotado aguardando login!")
                    _popup("Login nao foi concluido em 3 minutos.\nVerifique o navegador e tente novamente.")
                    await context.close()
                    return

                # Aguarda pagina carregar (com timeout curto para nao travar)
                try:
                    await page.wait_for_load_state("networkidle", timeout=10000)
                except Exception:
                    pass  # Continua mesmo se timeout
                await page.wait_for_timeout(2000)

                if "login" in page.url.lower():
                    print("  [ERRO] Login falhou!")
                    _popup("Login falhou. Verifique email/senha e tente novamente.")
                    await context.close()
                    return

            print("  [OK] Logado!\n")

            # RELATORIO 1: LISTA DE ANUNCIOS
            print("[RELATORIO] Abrindo Lista de Anuncios...")
            try:
                await page.goto(URL_ANUNCIOS, wait_until="domcontentloaded", timeout=30000)
            except Exception:
                pass
            await page.wait_for_timeout(2000)

            print("  [CONFIG] Selecionando Anuncios Ativos...")
            await page.evaluate("""
                var selects = document.querySelectorAll('select');
                for (var s of selects) {
                    for (var o of s.options) {
                        if (o.text.includes('Ativos')) {
                            s.value = o.value;
                            s.dispatchEvent(new Event('change'));
                            break;
                        }
                    }
                }
            """)
            await page.wait_for_timeout(500)

            print("  [BUSCANDO] Clicando em Buscar...")
            await page.click("#btnBuscar")
            await page.wait_for_function("""
                () => {
                    var btn = document.querySelector('#btnExportar');
                    return btn && !btn.classList.contains('disabled');
                }
            """, timeout=60000)
            print("  [OK] Resultados carregados!")

            print("  [DOWNLOAD] Exportando Excel...")
            async with page.expect_download(timeout=1800000) as dl:
                await page.evaluate("document.querySelector('#btnExportar').click()")
            download = await dl.value
            await download.save_as(str(PASTA_SAIDA / "listas_de_anuncios.xlsx"))
            print("  [OK] listas_de_anuncios.xlsx salvo!\n")

            # RELATORIO 2: MLB x ID INTERNO
            print("[RELATORIO] Abrindo MLB x ID Interno (SKU)...")
            try:
                await page.goto(URL_MLB_SKU, wait_until="domcontentloaded", timeout=30000)
            except Exception:
                pass
            await page.wait_for_timeout(2000)

            print("  [DOWNLOAD] Clicando em Exportar Lista de Produtos...")
            async with page.expect_download(timeout=1800000) as dl2:
                await page.evaluate("document.querySelector('#btnExportar').click()")
            download2 = await dl2.value
            await download2.save_as(str(PASTA_SAIDA / "mlb_id_interno.xlsx"))
            print("  [OK] mlb_id_interno.xlsx salvo!\n")

            await context.close()

    finally:
        shutil.rmtree(str(tmp_dir), ignore_errors=True)

    # PROCV COMPLETO: SKU + CUSTO + MARGEM
    print("[PROCV] Fazendo cruzamentos (SKU, Custo, Margem)...")
    fazer_procv_completo()

    print(f"\n[OK] TUDO PRONTO - Loja: {NOME_LOJA}")
    print(f"[OK] Arquivos em: {{PASTA_SAIDA.resolve()}}")
    print(f"[OK] Resultado final: listas_de_anuncios_COM_SKU_CUSTO_MARGEM.xlsx\n")
    _popup(f"Processo concluido!\nLoja: {NOME_LOJA}\nArquivo: listas_de_anuncios_COM_SKU_CUSTO_MARGEM.xlsx")


if __name__ == "__main__":
    asyncio.run(main())
